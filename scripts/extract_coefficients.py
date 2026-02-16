#!/usr/bin/env python3
"""Extract INTJ + Capricorn coefficient values from personality_building_blocks.xlsx.

SUPERSEDED by src/animus/data_pipeline/ in Phase 2. Kept for reference.

Reads the raw axis contributions (Sheet 1) and astrology components (Sheet 2),
computes the INTJ and Capricorn sums, validates against the pre-computed reference
values in Sheets 3 and 4, and prints values formatted for hardcoding.
"""

import sys
from pathlib import Path

import openpyxl

EXCEL_PATH = Path(__file__).parent.parent / "docs" / "personality_building_blocks.xlsx"

# MBTI axis pole positions in Sheet 1 (0-indexed row offsets within each coefficient block)
# Each block has: header, description, column headers, then 8 data rows (E, I, S, N, T, F, J, P)
MBTI_POLES = {
    "E": 0, "I": 1, "S": 2, "N": 3,
    "T": 4, "F": 5, "J": 6, "P": 7,
}

INTJ_POLES = ["I", "N", "T", "J"]

# Capricorn = Earth + Cardinal + Capricorn tweak
CAPRICORN_ELEMENT = "Earth"
CAPRICORN_MODALITY = "Cardinal"
CAPRICORN_SIGN = "Capricorn"


def parse_mbti_axes(ws) -> dict[str, dict[str, float]]:
    """Parse Sheet 1 into {coefficient_name: {pole: value}}."""
    coefficients = {}
    rows = list(ws.iter_rows(values_only=True))

    # Find coefficient blocks by scanning for known headers
    coeff_names = ["ASSERTIVENESS", "EMOTIONAL SUSCEPTIBILITY", "RIGIDITY", "RUMINATION"]
    appraisal_name = "APPRAISAL BASELINES"

    i = 0
    while i < len(rows):
        cell = rows[i][0]
        if cell in coeff_names:
            # Skip: header row, description row, column header row
            data_start = i + 3
            poles = {}
            for offset in range(8):
                row = rows[data_start + offset]
                pole_name = row[1]  # Pole name is in column B
                value = row[2]  # Value is in column C
                # Extract just the letter code from pole name
                code = pole_name.split("(")[1].rstrip(")") if "(" in pole_name else pole_name[0]
                poles[code] = float(value)
            coefficients[cell] = poles
        elif cell == appraisal_name:
            # Appraisal has two value columns: Control (C) and Certainty (D)
            data_start = i + 3
            control_poles = {}
            certainty_poles = {}
            for offset in range(8):
                row = rows[data_start + offset]
                pole_name = row[1]
                code = pole_name.split("(")[1].rstrip(")") if "(" in pole_name else pole_name[0]
                control_poles[code] = float(row[2])
                certainty_poles[code] = float(row[3])
            coefficients["CONTROL"] = control_poles
            coefficients["CERTAINTY"] = certainty_poles
        i += 1

    return coefficients


def parse_astrology(ws) -> tuple[dict, dict, dict]:
    """Parse Sheet 2 into elements, modalities, and sign tweaks."""
    rows = list(ws.iter_rows(values_only=True))
    elements = {}
    modalities = {}
    tweaks = {}

    section = None
    for row in rows:
        if row[0] == "ELEMENTS":
            section = "elements"
            continue
        elif row[0] == "MODALITIES":
            section = "modalities"
            continue
        elif row[0] and "SIGN-SPECIFIC TWEAKS" in str(row[0]):
            section = "tweaks"
            continue

        # Skip headers, column labels, and empty rows
        if row[0] is None:
            continue
        if row[0] in ("Element", "Modality", "Sign"):
            continue
        if all(v is None for v in row):
            continue
        # Skip description rows (no numeric data in column C)
        if row[2] is None:
            continue

        name = row[0]
        if section == "elements" and name in ("Fire", "Earth", "Air", "Water"):
            elements[name] = {
                "assertiveness": float(row[2]),
                "susceptibility": float(row[3]),
                "rigidity": float(row[4]),
                "rumination": float(row[5]),
                "control": float(row[6]),
                "certainty": float(row[7]),
            }
        elif section == "modalities" and name in ("Cardinal", "Fixed", "Mutable"):
            modalities[name] = {
                "assertiveness": float(row[2]),
                "susceptibility": float(row[3]),
                "rigidity": float(row[4]),
                "rumination": float(row[5]),
                "control": float(row[6]),
                "certainty": float(row[7]),
            }
        elif section == "tweaks":
            tweaks[name] = {
                "assertiveness": float(row[2]),
                "susceptibility": float(row[3]),
                "rigidity": float(row[4]),
                "rumination": float(row[5]),
                "control": float(row[6]),
                "certainty": float(row[7]),
            }

    return elements, modalities, tweaks


def compute_intj(mbti_axes: dict) -> dict[str, float]:
    """Sum the I + N + T + J poles for each coefficient."""
    result = {}
    coeff_map = {
        "ASSERTIVENESS": "assertiveness",
        "EMOTIONAL SUSCEPTIBILITY": "susceptibility",
        "RIGIDITY": "rigidity",
        "RUMINATION": "rumination",
        "CONTROL": "control",
        "CERTAINTY": "certainty",
    }
    for raw_name, clean_name in coeff_map.items():
        poles = mbti_axes[raw_name]
        result[clean_name] = sum(poles[p] for p in INTJ_POLES)
    return result


def compute_capricorn(elements, modalities, tweaks) -> dict[str, float]:
    """Sum element + modality + sign tweak for Capricorn."""
    result = {}
    for key in ("assertiveness", "susceptibility", "rigidity", "rumination", "control", "certainty"):
        result[key] = (
            elements[CAPRICORN_ELEMENT][key]
            + modalities[CAPRICORN_MODALITY][key]
            + tweaks[CAPRICORN_SIGN][key]
        )
    return result


def validate_against_sheet(ws, target_name: str, computed: dict[str, float], sheet_label: str) -> bool:
    """Validate computed values against reference row in Sheet 3 or 4."""
    rows = list(ws.iter_rows(values_only=True))

    # Detect column layout from header row
    headers = None
    for row in rows:
        if row[0] in ("Type", "Sign"):
            headers = [str(c) for c in row if c is not None]
            break

    for row in rows:
        if row[0] == target_name:
            # Find the first numeric column
            data_start = next(i for i, v in enumerate(row) if isinstance(v, (int, float)))
            ref = {
                "assertiveness": float(row[data_start]),
                "susceptibility": float(row[data_start + 1]),
                "rigidity": float(row[data_start + 2]),
                "rumination": float(row[data_start + 3]),
                "control": float(row[data_start + 4]),
                "certainty": float(row[data_start + 5]),
            }
            ok = True
            for key in ref:
                if abs(computed[key] - ref[key]) > 0.001:
                    print(f"  MISMATCH in {sheet_label} {target_name}.{key}: "
                          f"computed={computed[key]:.3f}, reference={ref[key]:.3f}")
                    ok = False
            if ok:
                print(f"  {sheet_label} {target_name}: all values match reference.")
            return ok
    print(f"  WARNING: {target_name} not found in {sheet_label}")
    return False


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # Parse source data
    mbti_axes = parse_mbti_axes(wb["MBTI Axis Contributions"])
    elements, modalities, tweaks = parse_astrology(wb["Astrology Components"])

    # Compute INTJ
    intj = compute_intj(mbti_axes)
    print("=== INTJ (pre-normalization sums) ===")
    for k, v in intj.items():
        print(f"  {k}: {v:.2f}")

    # Compute Capricorn
    capricorn = compute_capricorn(elements, modalities, tweaks)
    print("\n=== Capricorn (pre-normalization sums) ===")
    for k, v in capricorn.items():
        print(f"  {k}: {v:.2f}")

    # Validate
    print("\n=== Validation ===")
    validate_against_sheet(wb["16 MBTI Types (Computed)"], "INTJ", intj, "Sheet 3")
    validate_against_sheet(wb["12 Signs (Computed)"], "Capricorn", capricorn, "Sheet 4")

    # Print formatted for hardcoding
    print("\n=== Formatted for hardcoding ===")
    print(f"""
# INTJ coefficients (pre-normalization sums)
INTJ_ASSERTIVENESS = {intj['assertiveness']:.2f}
INTJ_SUSCEPTIBILITY = {intj['susceptibility']:.2f}
INTJ_RIGIDITY = {intj['rigidity']:.2f}
INTJ_RUMINATION = {intj['rumination']:.2f}
INTJ_CONTROL = {intj['control']:.2f}
INTJ_CERTAINTY = {intj['certainty']:.2f}

# Capricorn coefficients (pre-normalization sums)
CAP_ASSERTIVENESS = {capricorn['assertiveness']:.2f}
CAP_SUSCEPTIBILITY = {capricorn['susceptibility']:.2f}
CAP_RIGIDITY = {capricorn['rigidity']:.2f}
CAP_RUMINATION = {capricorn['rumination']:.2f}
CAP_CONTROL = {capricorn['control']:.2f}
CAP_CERTAINTY = {capricorn['certainty']:.2f}
""")


if __name__ == "__main__":
    main()
