#!/usr/bin/env python3
"""Populate the Excel workbook with transformation matrix, behavioral baseline,
and resting mood component data for all MBTI axis poles and astrology components.

This script adds new sheets to personality_building_blocks.xlsx:
- "MBTI Matrix Components": 5x5 matrix contributions per axis pole (E,I,S,N,T,F,J,P)
- "MBTI Baseline Components": behavioral baseline + resting mood contributions per pole
- "Astro Matrix Components": 5x5 matrix contributions per element/modality/sign tweak
- "Astro Baseline Components": baseline + mood contributions per element/modality/sign tweak
- Updates Sheets 3-4 with computed reference values for matrices, baselines, moods

Design rationale:
Each axis pole contributes a small 5x5 matrix that gets summed with 3 others to form
the MBTI type's transformation matrix. The matrix encodes how mood dimensions map to
behavioral dimensions.

Row labels: aggression, impulsiveness, sociability, empathy, curiosity
Col labels: distress_content, fear_confid, isol_belong, shame_pride, arousal
"""

import copy
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

EXCEL_PATH = Path(__file__).parent.parent / "docs" / "personality_building_blocks.xlsx"

BEHAVIORAL_DIMS = ["aggression", "impulsiveness", "sociability", "empathy", "curiosity"]
MOOD_DIMS = ["distress_content", "fear_confid", "isol_belong", "shame_pride", "arousal"]

# ============================================================================
# MBTI AXIS POLE MATRIX CONTRIBUTIONS
# ============================================================================
# Each pole contributes to the 5x5 transformation matrix.
# The type's matrix = sum of its 4 poles (e.g., INTJ = I + N + T + J).
#
# Design principles:
# - E/I axis: E amplifies sociability pathways, I dampens them
# - S/N axis: S amplifies direct/concrete responses, N amplifies curiosity
# - T/F axis: T dampens susceptibility overall (smaller values),
#              F amplifies emotional-to-behavioral translation
# - J/P axis: J amplifies deliberation, P amplifies impulsiveness
#
# Values are intentionally small because 4 poles sum together.
# Target range per cell after summing 4 poles: roughly -0.30 to +0.30

MBTI_POLE_MATRICES = {
    # Extraversion: socially activating, outwardly responsive
    "E": [
        #  dist    fear    isol    shame   arou
        [-0.02, -0.03,  0.00,  0.00,  0.05],   # aggression: arousal → slight aggression
        [-0.02, -0.02,  0.00,  0.00, -0.03],   # impulsiveness: slightly more reactive
        [-0.10, -0.06, -0.20,  0.02, -0.10],   # sociability: strong social seeking, esp from isolation
        [ 0.00,  0.00,  0.02,  0.02,  0.00],   # empathy: slightly more other-oriented
        [-0.02, -0.02, -0.02,  0.00, -0.05],   # curiosity: arousal drives engagement
    ],
    # Introversion: internally processing, socially withdrawing
    "I": [
        [-0.01, -0.02,  0.00, -0.02,  0.02],   # aggression: minimal
        [ 0.02,  0.03,  0.00,  0.02, -0.01],   # impulsiveness: more deliberate under stress
        [ 0.10,  0.10,  0.20,  0.05,  0.04],   # sociability: withdraws under stress / isolation
        [ 0.02,  0.02,  0.02,  0.03,  0.00],   # empathy: self-focused under stress
        [ 0.02,  0.02, -0.02,  0.02, -0.05],   # curiosity: inner curiosity from arousal
    ],
    # Sensing: grounded, concrete, present-focused
    "S": [
        [-0.03, -0.05, -0.02, -0.01,  0.05],   # aggression: fear→confrontation (concrete threats)
        [-0.02, -0.03,  0.00,  0.00, -0.05],   # impulsiveness: react to immediate situation
        [-0.01,  0.00, -0.05,  0.02, -0.02],   # sociability: practical social coping
        [ 0.01,  0.02,  0.00,  0.00,  0.03],   # empathy: mildly self-interested (practical focus)
        [-0.01, -0.01, -0.01,  0.00, -0.03],   # curiosity: arousal → investigate (tangible)
    ],
    # Intuition: abstract, vision-driven, pattern-seeking
    "N": [
        [-0.01, -0.01,  0.00, -0.01,  0.02],   # aggression: less direct, more strategic
        [ 0.01,  0.02,  0.00,  0.02, -0.02],   # impulsiveness: more deliberate (thinking ahead)
        [ 0.01,  0.00, -0.02,  0.01, -0.01],   # sociability: mild variation
        [ 0.01,  0.01,  0.01,  0.02,  0.00],   # empathy: slightly more other-aware (abstract)
        [ 0.03,  0.03, -0.03,  0.02, -0.08],   # curiosity: strong arousal→curiosity, explores patterns
    ],
    # Thinking: analytical, emotion-dampened, principled
    "T": [
        [-0.05, -0.07,  0.01, -0.01,  0.05],   # aggression: controlled pushback under fear
        [ 0.05,  0.06,  0.00,  0.04, -0.02],   # impulsiveness: deliberate, logic before action
        [ 0.03,  0.04, -0.01,  0.04,  0.01],   # sociability: withdraws to think
        [ 0.02,  0.02,  0.02,  0.04,  0.00],   # empathy: self-interested (logic > feeling)
        [ 0.02,  0.03, -0.01,  0.01, -0.05],   # curiosity: analytical curiosity
    ],
    # Feeling: emotion-amplified, values-driven, empathetic
    "F": [
        [-0.03, -0.04,  0.00, -0.04,  0.03],   # aggression: emotion→aggression (values threatened)
        [-0.03, -0.04,  0.00, -0.02, -0.05],   # impulsiveness: emotion drives immediate response
        [-0.04, -0.03, -0.08,  0.01, -0.03],   # sociability: seeks connection under stress
        [-0.02, -0.02, -0.01, -0.02, -0.02],   # empathy: more other-oriented (core F trait)
        [ 0.00,  0.00, -0.02, -0.01, -0.03],   # curiosity: less curiosity drive
    ],
    # Judging: structured, decisive, closure-seeking
    "J": [
        [-0.04, -0.05,  0.01, -0.01,  0.03],   # aggression: structured pushback
        [ 0.05,  0.06,  0.00,  0.04, -0.02],   # impulsiveness: strongly deliberate (core J trait)
        [ 0.02,  0.02, -0.01,  0.04,  0.02],   # sociability: maintains position
        [ 0.01,  0.01,  0.01,  0.02,  0.01],   # empathy: structured empathy
        [ 0.02,  0.02, -0.01,  0.01, -0.05],   # curiosity: focused, structured exploration
    ],
    # Perceiving: flexible, spontaneous, open
    "P": [
        [-0.04, -0.06, -0.02, -0.02,  0.08],   # aggression: arousal→impulsive aggression
        [-0.05, -0.06,  0.00, -0.02, -0.08],   # impulsiveness: strongly impulsive (core P trait)
        [-0.03, -0.03, -0.06,  0.01, -0.04],   # sociability: socially open, seeks stimulation
        [ 0.01,  0.02,  0.00,  0.00,  0.02],   # empathy: variable
        [-0.02, -0.02, -0.02,  0.00, -0.06],   # curiosity: strongly driven by novelty/arousal
    ],
}


# ============================================================================
# MBTI AXIS POLE BEHAVIORAL BASELINE CONTRIBUTIONS
# ============================================================================
# Each pole contributes to the 5D behavioral baseline.
# Dims: aggression, impulsiveness, sociability, empathy, curiosity

MBTI_POLE_BEHAVIORAL_BASELINES = {
    "E": [ 0.05,  0.05,  0.32,  0.00,  0.03],
    "I": [-0.05, -0.05, -0.32,  0.03,  0.02],
    "S": [ 0.03,  0.05, -0.02, -0.02,  0.00],
    "N": [-0.03, -0.05,  0.02,  0.02,  0.08],
    "T": [ 0.02, -0.10, -0.05,  0.08,  0.02],
    "F": [-0.02,  0.05,  0.05, -0.08,  0.00],
    "J": [ 0.00, -0.15,  0.00,  0.02,  0.00],
    "P": [ 0.05,  0.15,  0.02, -0.03,  0.03],
}


# ============================================================================
# MBTI AXIS POLE RESTING MOOD CONTRIBUTIONS
# ============================================================================
# Dims: distress_content, fear_confid, isol_belong, shame_pride, arousal

MBTI_POLE_RESTING_MOODS = {
    "E": [ 0.05,  0.08,  0.15,  0.02,  0.18],
    "I": [ 0.03,  0.02, -0.15, -0.01,  0.03],
    "S": [ 0.05,  0.05,  0.02,  0.02,  0.08],
    "N": [ 0.00,  0.03, -0.02,  0.01,  0.05],
    "T": [ 0.02,  0.08,  0.00,  0.05,  0.05],
    "F": [ 0.02, -0.02,  0.05, -0.02,  0.08],
    "J": [ 0.05,  0.05,  0.00,  0.03,  0.05],
    "P": [-0.02, -0.02,  0.02, -0.02,  0.10],
}


# ============================================================================
# ASTROLOGY COMPONENT MATRIX CONTRIBUTIONS
# ============================================================================
# Elements (4), Modalities (3), and Sign tweaks (12) each contribute
# to the 5x5 transformation matrix.
# Sign total = element matrix + modality matrix + tweak matrix

ASTRO_ELEMENT_MATRICES = {
    # Fire: bold, passionate, action-oriented. Strong arousal→aggression pathway.
    "Fire": [
        [-0.08, -0.10, -0.03, -0.02,  0.12],
        [-0.05, -0.06,  0.00,  0.00, -0.10],
        [-0.03, -0.02, -0.08,  0.03, -0.05],
        [ 0.01,  0.03,  0.00,  0.00,  0.03],
        [-0.03, -0.03, -0.03,  0.00, -0.08],
    ],
    # Earth: grounded, steady. Small matrix values (low susceptibility).
    "Earth": [
        [-0.02, -0.03,  0.00, -0.01,  0.02],
        [ 0.02,  0.02,  0.00,  0.01, -0.01],
        [ 0.02,  0.02,  0.00,  0.02,  0.00],
        [ 0.01,  0.01,  0.01,  0.02,  0.00],
        [ 0.01,  0.01,  0.00,  0.01, -0.02],
    ],
    # Air: intellectual, communicative. Curiosity and sociability pathways.
    "Air": [
        [-0.02, -0.03,  0.00, -0.01,  0.03],
        [ 0.01,  0.02,  0.00,  0.01, -0.02],
        [-0.02, -0.01, -0.05,  0.01, -0.03],
        [ 0.00,  0.00,  0.01,  0.01,  0.00],
        [ 0.02,  0.02, -0.02,  0.01, -0.06],
    ],
    # Water: emotional, deep. Strong mood→behavior translation.
    "Water": [
        [-0.05, -0.06, -0.02, -0.04,  0.05],
        [-0.03, -0.04,  0.00, -0.01, -0.05],
        [-0.04, -0.03, -0.10,  0.02, -0.04],
        [-0.02, -0.02, -0.02, -0.03, -0.01],
        [-0.01, -0.01, -0.02, -0.01, -0.04],
    ],
}

ASTRO_MODALITY_MATRICES = {
    # Cardinal: initiators. Activating pathways.
    "Cardinal": [
        [-0.04, -0.05, -0.02, -0.01,  0.05],
        [-0.02, -0.03,  0.00,  0.00, -0.04],
        [-0.02, -0.01, -0.05,  0.02, -0.03],
        [ 0.01,  0.01,  0.01,  0.01,  0.01],
        [-0.01, -0.01, -0.01,  0.00, -0.04],
    ],
    # Fixed: persistent. Maintains state, resists change.
    "Fixed": [
        [-0.02, -0.03, -0.01, -0.01,  0.03],
        [ 0.02,  0.03,  0.00,  0.02, -0.02],
        [ 0.01,  0.01, -0.03,  0.02,  0.00],
        [ 0.01,  0.01,  0.01,  0.02,  0.00],
        [ 0.01,  0.01, -0.01,  0.01, -0.03],
    ],
    # Mutable: adaptable. Responsive, flexible pathways.
    "Mutable": [
        [-0.05, -0.06, -0.02, -0.02,  0.07],
        [-0.04, -0.05,  0.00, -0.01, -0.06],
        [-0.03, -0.02, -0.06,  0.01, -0.04],
        [ 0.00,  0.01,  0.00,  0.00,  0.01],
        [-0.02, -0.02, -0.02, -0.01, -0.06],
    ],
}

ASTRO_TWEAK_MATRICES = {
    # Most tweaks are zero or very small. Only significant where the sign
    # deviates meaningfully from element+modality prediction.
    "Aries": [
        [ 0.00,  0.00,  0.00,  0.00,  0.02],  # extra arousal→aggression
        [ 0.00,  0.00,  0.00,  0.00, -0.01],
        [ 0.00,  0.00,  0.00,  0.00,  0.00],
        [ 0.00,  0.00,  0.00,  0.00,  0.00],
        [ 0.00,  0.00,  0.00,  0.00, -0.01],
    ],
    "Taurus": [
        [ 0.00,  0.00,  0.00,  0.00,  0.00],
        [ 0.01,  0.01,  0.00,  0.00,  0.00],  # extra deliberation
        [ 0.00,  0.00,  0.00,  0.00,  0.00],
        [ 0.00,  0.00,  0.00,  0.00,  0.00],
        [ 0.00,  0.00,  0.00,  0.00,  0.00],
    ],
    "Gemini": [
        [ 0.00,  0.00,  0.00,  0.00,  0.00],
        [ 0.00,  0.00,  0.00,  0.00, -0.01],  # extra impulsiveness
        [ 0.00,  0.00, -0.01,  0.00, -0.01],  # extra social seeking
        [ 0.00,  0.00,  0.00,  0.00,  0.00],
        [ 0.01,  0.01, -0.01,  0.00, -0.02],  # extra curiosity
    ],
    "Cancer": [
        [ 0.00,  0.00,  0.00, -0.01,  0.00],
        [ 0.00,  0.00,  0.00,  0.00,  0.00],
        [-0.02, -0.01, -0.02,  0.00, -0.01],  # extra social seeking (nurturing)
        [-0.01, -0.01, -0.01, -0.01,  0.00],  # extra empathy
        [ 0.00,  0.00,  0.00,  0.00,  0.00],
    ],
    "Leo": [
        [-0.01, -0.01,  0.00,  0.00,  0.02],  # extra arousal→aggression (assertive)
        [ 0.00,  0.00,  0.00,  0.00,  0.00],
        [-0.01, -0.01, -0.01,  0.01, -0.01],  # social engagement
        [ 0.00,  0.00,  0.00,  0.00,  0.00],
        [ 0.00,  0.00,  0.00,  0.00, -0.01],
    ],
    "Virgo": [
        [ 0.00,  0.00,  0.00,  0.00,  0.00],
        [ 0.01,  0.02,  0.00,  0.01, -0.01],  # extra deliberation (overthinking)
        [ 0.01,  0.01,  0.00,  0.01,  0.00],  # slight withdrawal
        [ 0.00,  0.00,  0.01,  0.01,  0.00],
        [ 0.01,  0.01,  0.00,  0.00, -0.01],  # analytical curiosity
    ],
    "Libra": [
        [ 0.02,  0.02,  0.00,  0.00, -0.02],  # avoids aggression
        [ 0.00,  0.00,  0.00,  0.00,  0.00],
        [-0.01, -0.01, -0.01,  0.00, -0.01],  # seeks social connection
        [-0.01, -0.01,  0.00, -0.01,  0.00],  # more empathetic
        [ 0.00,  0.00,  0.00,  0.00,  0.00],
    ],
    "Scorpio": [
        [-0.02, -0.02, -0.01, -0.02,  0.02],  # intensified aggression
        [ 0.00,  0.00,  0.00,  0.00,  0.00],
        [ 0.02,  0.02,  0.00,  0.02,  0.01],  # less social (secretive)
        [ 0.00,  0.00,  0.00,  0.00,  0.00],
        [-0.01, -0.01, -0.01, -0.01, -0.01],  # investigative curiosity
    ],
    "Sagittarius": [
        [ 0.00,  0.00,  0.00,  0.00,  0.01],
        [-0.01, -0.01,  0.00,  0.00, -0.01],  # extra impulsiveness
        [-0.01,  0.00, -0.01,  0.00, -0.01],
        [ 0.00,  0.00,  0.00,  0.00,  0.00],
        [-0.01, -0.01, -0.01,  0.00, -0.02],  # extra curiosity (adventurous)
    ],
    "Capricorn": [
        [ 0.00,  0.00,  0.00,  0.00,  0.00],
        [ 0.01,  0.01,  0.00,  0.01,  0.00],  # extra deliberation
        [ 0.01,  0.01,  0.00,  0.01,  0.00],  # slight withdrawal
        [ 0.00,  0.00,  0.00,  0.00,  0.00],
        [ 0.00,  0.00,  0.00,  0.00, -0.01],
    ],
    "Aquarius": [
        [ 0.00,  0.00,  0.00,  0.00,  0.00],
        [ 0.00,  0.01,  0.00,  0.00,  0.00],
        [ 0.01,  0.01,  0.00,  0.01,  0.00],  # slight withdrawal (detached)
        [ 0.00,  0.00,  0.00,  0.00,  0.00],
        [ 0.01,  0.01, -0.01,  0.00, -0.01],  # intellectual curiosity
    ],
    "Pisces": [
        [ 0.02,  0.02,  0.00,  0.00, -0.01],  # less aggression
        [-0.01, -0.01,  0.00,  0.00, -0.01],  # more impulsive (swept along)
        [-0.02, -0.02, -0.02,  0.00, -0.01],  # social seeking
        [-0.02, -0.02, -0.01, -0.02, -0.01],  # strong empathy
        [ 0.00,  0.00, -0.01,  0.00, -0.01],  # mild curiosity
    ],
}


# ============================================================================
# ASTROLOGY BEHAVIORAL BASELINE CONTRIBUTIONS
# ============================================================================
# Dims: aggression, impulsiveness, sociability, empathy, curiosity

ASTRO_ELEMENT_BEHAVIORAL_BASELINES = {
    "Fire":  [ 0.10,  0.10,  0.05, -0.02,  0.08],
    "Earth": [-0.05, -0.08, -0.05,  0.03, -0.02],
    "Air":   [ 0.00,  0.00,  0.08,  0.00,  0.08],
    "Water": [-0.05, -0.02,  0.03, -0.08,  0.00],
}

ASTRO_MODALITY_BEHAVIORAL_BASELINES = {
    "Cardinal": [ 0.05,  0.02,  0.02,  0.00,  0.03],
    "Fixed":    [ 0.00, -0.05, -0.02,  0.02, -0.02],
    "Mutable":  [-0.02,  0.05,  0.03, -0.01,  0.05],
}

ASTRO_TWEAK_BEHAVIORAL_BASELINES = {
    "Aries":       [ 0.05,  0.05,  0.00,  0.00,  0.02],
    "Taurus":      [ 0.00, -0.02,  0.00,  0.02,  0.00],
    "Gemini":      [ 0.00,  0.03,  0.02,  0.00,  0.03],
    "Cancer":      [ 0.00,  0.00,  0.05, -0.05,  0.00],
    "Leo":         [ 0.05,  0.00,  0.05,  0.00,  0.02],
    "Virgo":       [-0.02, -0.03, -0.02,  0.02,  0.02],
    "Libra":       [-0.05,  0.00,  0.05, -0.03,  0.00],
    "Scorpio":     [ 0.05,  0.00, -0.05,  0.00,  0.03],
    "Sagittarius": [ 0.02,  0.03,  0.02,  0.00,  0.05],
    "Capricorn":   [-0.02, -0.03, -0.05,  0.03, -0.02],
    "Aquarius":    [ 0.00, -0.02, -0.03,  0.00,  0.03],
    "Pisces":      [-0.05,  0.00,  0.03, -0.05, -0.02],
}


# ============================================================================
# ASTROLOGY RESTING MOOD CONTRIBUTIONS
# ============================================================================
# Dims: distress_content, fear_confid, isol_belong, shame_pride, arousal

ASTRO_ELEMENT_RESTING_MOODS = {
    "Fire":  [ 0.05,  0.10,  0.05,  0.03,  0.15],
    "Earth": [ 0.08,  0.05, -0.02,  0.05,  0.05],
    "Air":   [ 0.03,  0.05,  0.05,  0.02,  0.08],
    "Water": [-0.02, -0.05,  0.00, -0.03,  0.05],
}

ASTRO_MODALITY_RESTING_MOODS = {
    "Cardinal": [ 0.03,  0.05,  0.02,  0.02,  0.08],
    "Fixed":    [ 0.05,  0.05, -0.02,  0.03,  0.05],
    "Mutable":  [ 0.00,  0.00,  0.03, -0.01,  0.08],
}

ASTRO_TWEAK_RESTING_MOODS = {
    "Aries":       [ 0.02,  0.03,  0.00,  0.00,  0.02],
    "Taurus":      [ 0.02,  0.02,  0.00,  0.02,  0.00],
    "Gemini":      [ 0.00,  0.00,  0.03, -0.01,  0.02],
    "Cancer":      [-0.03, -0.02,  0.02, -0.02,  0.00],
    "Leo":         [ 0.03,  0.05,  0.02,  0.05,  0.02],
    "Virgo":       [-0.02, -0.03, -0.02, -0.02,  0.02],
    "Libra":       [ 0.02,  0.00,  0.03,  0.00,  0.00],
    "Scorpio":     [-0.02, -0.03, -0.03, -0.02,  0.02],
    "Sagittarius": [ 0.03,  0.03,  0.02,  0.02,  0.02],
    "Capricorn":   [ 0.02,  0.02, -0.03,  0.03, -0.02],
    "Aquarius":    [ 0.00,  0.00, -0.02,  0.00,  0.00],
    "Pisces":      [-0.03, -0.03,  0.02, -0.03,  0.00],
}


# ============================================================================
# MBTI TYPE POLE COMPOSITIONS (for reference/validation)
# ============================================================================
MBTI_TYPE_POLES = {
    "INTJ": ("I", "N", "T", "J"),
    "ENTJ": ("E", "N", "T", "J"),
    "INTP": ("I", "N", "T", "P"),
    "ENTP": ("E", "N", "T", "P"),
    "INFJ": ("I", "N", "F", "J"),
    "ENFJ": ("E", "N", "F", "J"),
    "INFP": ("I", "N", "F", "P"),
    "ENFP": ("E", "N", "F", "P"),
    "ISTJ": ("I", "S", "T", "J"),
    "ESTJ": ("E", "S", "T", "J"),
    "ISTP": ("I", "S", "T", "P"),
    "ESTP": ("E", "S", "T", "P"),
    "ISFJ": ("I", "S", "F", "J"),
    "ESFJ": ("E", "S", "F", "J"),
    "ISFP": ("I", "S", "F", "P"),
    "ESFP": ("E", "S", "F", "P"),
}

SIGN_COMPONENTS = {
    "Aries": ("Fire", "Cardinal"),
    "Taurus": ("Earth", "Fixed"),
    "Gemini": ("Air", "Mutable"),
    "Cancer": ("Water", "Cardinal"),
    "Leo": ("Fire", "Fixed"),
    "Virgo": ("Earth", "Mutable"),
    "Libra": ("Air", "Cardinal"),
    "Scorpio": ("Water", "Fixed"),
    "Sagittarius": ("Fire", "Mutable"),
    "Capricorn": ("Earth", "Cardinal"),
    "Aquarius": ("Air", "Fixed"),
    "Pisces": ("Water", "Mutable"),
}


def sum_matrices(matrices: list[list[list[float]]]) -> list[list[float]]:
    """Sum a list of 5x5 matrices element-wise."""
    result = [[0.0]*5 for _ in range(5)]
    for m in matrices:
        for i in range(5):
            for j in range(5):
                result[i][j] += m[i][j]
    return result


def sum_vectors(vectors: list[list[float]]) -> list[float]:
    """Sum a list of vectors element-wise."""
    n = len(vectors[0])
    return [sum(v[i] for v in vectors) for i in range(n)]


def compute_mbti_type_matrix(type_name: str) -> list[list[float]]:
    poles = MBTI_TYPE_POLES[type_name]
    return sum_matrices([MBTI_POLE_MATRICES[p] for p in poles])


def compute_mbti_type_baseline(type_name: str) -> list[float]:
    poles = MBTI_TYPE_POLES[type_name]
    return sum_vectors([MBTI_POLE_BEHAVIORAL_BASELINES[p] for p in poles])


def compute_mbti_type_mood(type_name: str) -> list[float]:
    poles = MBTI_TYPE_POLES[type_name]
    return sum_vectors([MBTI_POLE_RESTING_MOODS[p] for p in poles])


def compute_sign_matrix(sign: str) -> list[list[float]]:
    element, modality = SIGN_COMPONENTS[sign]
    return sum_matrices([
        ASTRO_ELEMENT_MATRICES[element],
        ASTRO_MODALITY_MATRICES[modality],
        ASTRO_TWEAK_MATRICES[sign],
    ])


def compute_sign_baseline(sign: str) -> list[float]:
    element, modality = SIGN_COMPONENTS[sign]
    return sum_vectors([
        ASTRO_ELEMENT_BEHAVIORAL_BASELINES[element],
        ASTRO_MODALITY_BEHAVIORAL_BASELINES[modality],
        ASTRO_TWEAK_BEHAVIORAL_BASELINES[sign],
    ])


def compute_sign_mood(sign: str) -> list[float]:
    element, modality = SIGN_COMPONENTS[sign]
    return sum_vectors([
        ASTRO_ELEMENT_RESTING_MOODS[element],
        ASTRO_MODALITY_RESTING_MOODS[modality],
        ASTRO_TWEAK_RESTING_MOODS[sign],
    ])


def write_mbti_matrix_sheet(wb):
    """Write MBTI axis pole transformation matrix contributions."""
    if "MBTI Matrix Components" in wb.sheetnames:
        del wb["MBTI Matrix Components"]
    ws = wb.create_sheet("MBTI Matrix Components")

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    row = 1
    ws.cell(row=row, column=1, value="MBTI Axis Pole Transformation Matrix Contributions").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value="Each pole contributes a 5x5 matrix. A type's matrix = sum of its 4 axis poles.")
    row += 2

    poles_order = ["E", "I", "S", "N", "T", "F", "J", "P"]
    pole_names = {
        "E": "Extraversion (E)", "I": "Introversion (I)",
        "S": "Sensing (S)", "N": "Intuition (N)",
        "T": "Thinking (T)", "F": "Feeling (F)",
        "J": "Judging (J)", "P": "Perceiving (P)",
    }

    for pole in poles_order:
        ws.cell(row=row, column=1, value=pole_names[pole]).font = header_font
        row += 1
        # Column headers
        ws.cell(row=row, column=1, value="Behavioral \\ Mood")
        for j, mood in enumerate(MOOD_DIMS):
            ws.cell(row=row, column=2+j, value=mood).font = header_font
        row += 1
        # Matrix data
        matrix = MBTI_POLE_MATRICES[pole]
        for i, beh in enumerate(BEHAVIORAL_DIMS):
            ws.cell(row=row, column=1, value=beh)
            for j in range(5):
                ws.cell(row=row, column=2+j, value=matrix[i][j])
            row += 1
        row += 1  # blank line

    return ws


def write_mbti_baseline_sheet(wb):
    """Write MBTI axis pole behavioral baseline and resting mood contributions."""
    if "MBTI Baseline Components" in wb.sheetnames:
        del wb["MBTI Baseline Components"]
    ws = wb.create_sheet("MBTI Baseline Components")

    header_font = Font(bold=True)
    row = 1
    ws.cell(row=row, column=1, value="MBTI Axis Pole Behavioral Baseline & Resting Mood Contributions").font = Font(bold=True, size=12)
    row += 2

    # Behavioral baselines
    ws.cell(row=row, column=1, value="BEHAVIORAL BASELINES").font = header_font
    row += 1
    ws.cell(row=row, column=1, value="Pole")
    for j, dim in enumerate(BEHAVIORAL_DIMS):
        ws.cell(row=row, column=2+j, value=dim).font = header_font
    row += 1
    for pole in ["E", "I", "S", "N", "T", "F", "J", "P"]:
        ws.cell(row=row, column=1, value=pole)
        for j, val in enumerate(MBTI_POLE_BEHAVIORAL_BASELINES[pole]):
            ws.cell(row=row, column=2+j, value=val)
        row += 1

    row += 1
    # Resting moods
    ws.cell(row=row, column=1, value="RESTING MOODS").font = header_font
    row += 1
    ws.cell(row=row, column=1, value="Pole")
    for j, dim in enumerate(MOOD_DIMS):
        ws.cell(row=row, column=2+j, value=dim).font = header_font
    row += 1
    for pole in ["E", "I", "S", "N", "T", "F", "J", "P"]:
        ws.cell(row=row, column=1, value=pole)
        for j, val in enumerate(MBTI_POLE_RESTING_MOODS[pole]):
            ws.cell(row=row, column=2+j, value=val)
        row += 1

    return ws


def write_astro_matrix_sheet(wb):
    """Write astrology component transformation matrix contributions."""
    if "Astro Matrix Components" in wb.sheetnames:
        del wb["Astro Matrix Components"]
    ws = wb.create_sheet("Astro Matrix Components")

    header_font = Font(bold=True)
    row = 1
    ws.cell(row=row, column=1, value="Astrological Component Transformation Matrix Contributions").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value="Sign matrix = element matrix + modality matrix + sign tweak matrix.")
    row += 2

    def write_section(ws, row, title, data, keys):
        ws.cell(row=row, column=1, value=title).font = header_font
        row += 1
        for key in keys:
            ws.cell(row=row, column=1, value=key).font = Font(bold=True)
            row += 1
            ws.cell(row=row, column=1, value="Behavioral \\ Mood")
            for j, mood in enumerate(MOOD_DIMS):
                ws.cell(row=row, column=2+j, value=mood).font = header_font
            row += 1
            matrix = data[key]
            for i, beh in enumerate(BEHAVIORAL_DIMS):
                ws.cell(row=row, column=1, value=beh)
                for j in range(5):
                    ws.cell(row=row, column=2+j, value=matrix[i][j])
                row += 1
            row += 1
        return row

    row = write_section(ws, row, "ELEMENTS", ASTRO_ELEMENT_MATRICES,
                       ["Fire", "Earth", "Air", "Water"])
    row = write_section(ws, row, "MODALITIES", ASTRO_MODALITY_MATRICES,
                       ["Cardinal", "Fixed", "Mutable"])
    row = write_section(ws, row, "SIGN-SPECIFIC TWEAKS", ASTRO_TWEAK_MATRICES,
                       list(SIGN_COMPONENTS.keys()))

    return ws


def write_astro_baseline_sheet(wb):
    """Write astrology component baseline and resting mood contributions."""
    if "Astro Baseline Components" in wb.sheetnames:
        del wb["Astro Baseline Components"]
    ws = wb.create_sheet("Astro Baseline Components")

    header_font = Font(bold=True)
    row = 1
    ws.cell(row=row, column=1, value="Astrological Component Behavioral Baseline & Resting Mood Contributions").font = Font(bold=True, size=12)
    row += 2

    def write_vector_section(ws, row, title, data, keys, dim_names):
        ws.cell(row=row, column=1, value=title).font = header_font
        row += 1
        ws.cell(row=row, column=1, value="Component")
        for j, dim in enumerate(dim_names):
            ws.cell(row=row, column=2+j, value=dim).font = header_font
        row += 1
        for key in keys:
            ws.cell(row=row, column=1, value=key)
            for j, val in enumerate(data[key]):
                ws.cell(row=row, column=2+j, value=val)
            row += 1
        row += 1
        return row

    # Behavioral baselines
    ws.cell(row=row, column=1, value="BEHAVIORAL BASELINES").font = Font(bold=True, size=11)
    row += 1
    row = write_vector_section(ws, row, "Elements", ASTRO_ELEMENT_BEHAVIORAL_BASELINES,
                               ["Fire", "Earth", "Air", "Water"], BEHAVIORAL_DIMS)
    row = write_vector_section(ws, row, "Modalities", ASTRO_MODALITY_BEHAVIORAL_BASELINES,
                               ["Cardinal", "Fixed", "Mutable"], BEHAVIORAL_DIMS)
    row = write_vector_section(ws, row, "Sign Tweaks", ASTRO_TWEAK_BEHAVIORAL_BASELINES,
                               list(SIGN_COMPONENTS.keys()), BEHAVIORAL_DIMS)

    row += 1
    ws.cell(row=row, column=1, value="RESTING MOODS").font = Font(bold=True, size=11)
    row += 1
    row = write_vector_section(ws, row, "Elements", ASTRO_ELEMENT_RESTING_MOODS,
                               ["Fire", "Earth", "Air", "Water"], MOOD_DIMS)
    row = write_vector_section(ws, row, "Modalities", ASTRO_MODALITY_RESTING_MOODS,
                               ["Cardinal", "Fixed", "Mutable"], MOOD_DIMS)
    row = write_vector_section(ws, row, "Sign Tweaks", ASTRO_TWEAK_RESTING_MOODS,
                               list(SIGN_COMPONENTS.keys()), MOOD_DIMS)

    return ws


def update_mbti_reference_sheet(wb):
    """Update Sheet 3 with computed matrix, baseline, and mood reference values."""
    ws = wb["16 MBTI Types (Computed)"]

    # Find the current header row
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row[0] == "Type":
            header_row = i
            break

    if header_row is None:
        return

    # Current headers end at column 8. Add new columns starting at 9.
    col = 9
    header_font = Font(bold=True)

    # Add matrix headers (25 columns: 5 behavioral × 5 mood)
    for beh in BEHAVIORAL_DIMS:
        for mood in MOOD_DIMS:
            ws.cell(row=header_row, column=col, value=f"M:{beh[:3]}←{mood[:4]}").font = header_font
            col += 1

    # Add baseline headers (5 columns)
    for dim in BEHAVIORAL_DIMS:
        ws.cell(row=header_row, column=col, value=f"BL:{dim[:5]}").font = header_font
        col += 1

    # Add mood headers (5 columns)
    for dim in MOOD_DIMS:
        ws.cell(row=header_row, column=col, value=f"RM:{dim[:5]}").font = header_font
        col += 1

    # Fill in computed values for each type
    type_order = [
        "ISTJ", "ISFJ", "INFJ", "INTJ", "ISTP", "ISFP", "INFP", "INTP",
        "ESTP", "ESFP", "ENFP", "ENTP", "ESTJ", "ESFJ", "ENFJ", "ENTJ",
    ]

    for idx, type_name in enumerate(type_order):
        data_row = header_row + 1 + idx
        col = 9

        matrix = compute_mbti_type_matrix(type_name)
        for i in range(5):
            for j in range(5):
                ws.cell(row=data_row, column=col, value=round(matrix[i][j], 4))
                col += 1

        baseline = compute_mbti_type_baseline(type_name)
        for val in baseline:
            ws.cell(row=data_row, column=col, value=round(val, 4))
            col += 1

        mood = compute_mbti_type_mood(type_name)
        for val in mood:
            ws.cell(row=data_row, column=col, value=round(val, 4))
            col += 1


def update_sign_reference_sheet(wb):
    """Update Sheet 4 with computed matrix, baseline, and mood reference values."""
    ws = wb["12 Signs (Computed)"]

    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row[0] == "Sign":
            header_row = i
            break

    if header_row is None:
        return

    col = 11  # After existing columns
    header_font = Font(bold=True)

    for beh in BEHAVIORAL_DIMS:
        for mood in MOOD_DIMS:
            ws.cell(row=header_row, column=col, value=f"M:{beh[:3]}←{mood[:4]}").font = header_font
            col += 1

    for dim in BEHAVIORAL_DIMS:
        ws.cell(row=header_row, column=col, value=f"BL:{dim[:5]}").font = header_font
        col += 1

    for dim in MOOD_DIMS:
        ws.cell(row=header_row, column=col, value=f"RM:{dim[:5]}").font = header_font
        col += 1

    sign_order = [
        "Aries", "Taurus", "Gemini", "Cancer", "Leo", "Virgo",
        "Libra", "Scorpio", "Sagittarius", "Capricorn", "Aquarius", "Pisces",
    ]

    for idx, sign_name in enumerate(sign_order):
        data_row = header_row + 1 + idx
        col = 11

        matrix = compute_sign_matrix(sign_name)
        for i in range(5):
            for j in range(5):
                ws.cell(row=data_row, column=col, value=round(matrix[i][j], 4))
                col += 1

        baseline = compute_sign_baseline(sign_name)
        for val in baseline:
            ws.cell(row=data_row, column=col, value=round(val, 4))
            col += 1

        mood = compute_sign_mood(sign_name)
        for val in mood:
            ws.cell(row=data_row, column=col, value=round(val, 4))
            col += 1


def verify_intj_capricorn():
    """Print computed INTJ-Capricorn values for manual comparison with Phase 1."""
    print("=== INTJ Computed Matrix ===")
    m = compute_mbti_type_matrix("INTJ")
    for i, beh in enumerate(BEHAVIORAL_DIMS):
        print(f"  {beh}: {[round(v, 4) for v in m[i]]}")

    print("\n=== INTJ Behavioral Baseline ===")
    b = compute_mbti_type_baseline("INTJ")
    print(f"  {[round(v, 4) for v in b]}")

    print("\n=== INTJ Resting Mood ===")
    rm = compute_mbti_type_mood("INTJ")
    print(f"  {[round(v, 4) for v in rm]}")

    print("\n=== Capricorn Computed Matrix ===")
    m = compute_sign_matrix("Capricorn")
    for i, beh in enumerate(BEHAVIORAL_DIMS):
        print(f"  {beh}: {[round(v, 4) for v in m[i]]}")

    print("\n=== Capricorn Behavioral Baseline ===")
    b = compute_sign_baseline("Capricorn")
    print(f"  {[round(v, 4) for v in b]}")

    print("\n=== Capricorn Resting Mood ===")
    rm = compute_sign_mood("Capricorn")
    print(f"  {[round(v, 4) for v in rm]}")

    print("\n=== ESTP Computed Matrix ===")
    m = compute_mbti_type_matrix("ESTP")
    for i, beh in enumerate(BEHAVIORAL_DIMS):
        print(f"  {beh}: {[round(v, 4) for v in m[i]]}")

    print("\n=== Aries Computed Matrix ===")
    m = compute_sign_matrix("Aries")
    for i, beh in enumerate(BEHAVIORAL_DIMS):
        print(f"  {beh}: {[round(v, 4) for v in m[i]]}")


def main():
    verify_intj_capricorn()

    wb = openpyxl.load_workbook(EXCEL_PATH)

    write_mbti_matrix_sheet(wb)
    write_mbti_baseline_sheet(wb)
    write_astro_matrix_sheet(wb)
    write_astro_baseline_sheet(wb)
    update_mbti_reference_sheet(wb)
    update_sign_reference_sheet(wb)

    wb.save(EXCEL_PATH)
    print(f"\nExcel file updated: {EXCEL_PATH}")


if __name__ == "__main__":
    main()
