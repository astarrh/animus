"""Tier 1: Read raw coefficient data from Excel Sheets 1-2 and component sheets.

Returns structured raw data with no computation — just faithful extraction
of what the spreadsheet contains.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl

DEFAULT_EXCEL_PATH = Path(__file__).parent.parent.parent.parent / "docs" / "personality_building_blocks.xlsx"


# ---------------------------------------------------------------------------
# Raw data containers
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class RawMBTIAxes:
    """Raw axis pole values from Sheet 1.

    poles: {coefficient_name: {pole_letter: value}}
        e.g. {"assertiveness": {"E": 0.55, "I": 0.35, ...}, ...}
    Keys: assertiveness, susceptibility, rigidity, rumination, control, certainty
    """

    poles: dict[str, dict[str, float]]


@dataclass(frozen=True)
class RawMBTIMatrixComponents:
    """Raw transformation matrix contributions per MBTI axis pole.

    matrices: {pole_letter: list[list[float]]}  — 5x5 matrix per pole
    """

    matrices: dict[str, list[list[float]]]


@dataclass(frozen=True)
class RawMBTIBaselineComponents:
    """Raw behavioral baseline and resting mood contributions per MBTI axis pole.

    behavioral_baselines: {pole_letter: list[float]}  — 5 values per pole
    resting_moods: {pole_letter: list[float]}  — 5 values per pole
    """

    behavioral_baselines: dict[str, list[float]]
    resting_moods: dict[str, list[float]]


@dataclass(frozen=True)
class RawAstrologyComponents:
    """Raw scalar coefficient values from Sheet 2.

    Each dict maps component name to {coefficient: value}.
    """

    elements: dict[str, dict[str, float]]
    modalities: dict[str, dict[str, float]]
    tweaks: dict[str, dict[str, float]]


@dataclass(frozen=True)
class RawAstroMatrixComponents:
    """Raw transformation matrix contributions per astrology component.

    Each dict maps component name to a 5x5 matrix (list of lists).
    """

    elements: dict[str, list[list[float]]]
    modalities: dict[str, list[list[float]]]
    tweaks: dict[str, list[list[float]]]


@dataclass(frozen=True)
class RawAstroBaselineComponents:
    """Raw behavioral baseline and resting mood contributions per astrology component."""

    element_baselines: dict[str, list[float]]
    element_moods: dict[str, list[float]]
    modality_baselines: dict[str, list[float]]
    modality_moods: dict[str, list[float]]
    tweak_baselines: dict[str, list[float]]
    tweak_moods: dict[str, list[float]]


@dataclass(frozen=True)
class RawExcelData:
    """Complete raw data from all Excel sheets."""

    mbti_axes: RawMBTIAxes
    mbti_matrices: RawMBTIMatrixComponents
    mbti_baselines: RawMBTIBaselineComponents
    astrology: RawAstrologyComponents
    astro_matrices: RawAstroMatrixComponents
    astro_baselines: RawAstroBaselineComponents


# ---------------------------------------------------------------------------
# Parsing functions
# ---------------------------------------------------------------------------

def parse_excel(path: Path | None = None) -> RawExcelData:
    """Parse all sheets and return structured raw data."""
    if path is None:
        path = DEFAULT_EXCEL_PATH
    wb = openpyxl.load_workbook(path, data_only=True)

    mbti_axes = _parse_mbti_axes(wb["MBTI Axis Contributions"])
    astrology = _parse_astrology(wb["Astrology Components"])
    mbti_matrices = _parse_mbti_matrix_components(wb["MBTI Matrix Components"])
    mbti_baselines = _parse_mbti_baseline_components(wb["MBTI Baseline Components"])
    astro_matrices = _parse_astro_matrix_components(wb["Astro Matrix Components"])
    astro_baselines = _parse_astro_baseline_components(wb["Astro Baseline Components"])

    return RawExcelData(
        mbti_axes=mbti_axes,
        mbti_matrices=mbti_matrices,
        mbti_baselines=mbti_baselines,
        astrology=astrology,
        astro_matrices=astro_matrices,
        astro_baselines=astro_baselines,
    )


def parse_reference_mbti(path: Path | None = None) -> dict[str, dict[str, float]]:
    """Read Sheet 3 scalar reference values for all 16 MBTI types.

    Returns {type_name: {coefficient: value}} with pre-normalization sums.
    """
    if path is None:
        path = DEFAULT_EXCEL_PATH
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["16 MBTI Types (Computed)"]
    return _parse_reference_sheet(ws, "Type")


def parse_reference_signs(path: Path | None = None) -> dict[str, dict[str, float]]:
    """Read Sheet 4 scalar reference values for all 12 signs.

    Returns {sign_name: {coefficient: value}} with pre-normalization sums.
    """
    if path is None:
        path = DEFAULT_EXCEL_PATH
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["12 Signs (Computed)"]
    return _parse_reference_sheet(ws, "Sign")


# ---------------------------------------------------------------------------
# Internal parsing helpers
# ---------------------------------------------------------------------------

def _parse_mbti_axes(ws) -> RawMBTIAxes:
    """Parse Sheet 1 (MBTI Axis Contributions) into {coefficient: {pole: value}}."""
    coefficients: dict[str, dict[str, float]] = {}
    rows = list(ws.iter_rows(values_only=True))

    coeff_names = ["ASSERTIVENESS", "EMOTIONAL SUSCEPTIBILITY", "RIGIDITY", "RUMINATION"]
    coeff_map = {
        "ASSERTIVENESS": "assertiveness",
        "EMOTIONAL SUSCEPTIBILITY": "susceptibility",
        "RIGIDITY": "rigidity",
        "RUMINATION": "rumination",
    }

    i = 0
    while i < len(rows):
        cell = rows[i][0]
        if cell in coeff_names:
            data_start = i + 3  # skip header, description, column header
            poles: dict[str, float] = {}
            for offset in range(8):
                row = rows[data_start + offset]
                pole_name = row[1]
                value = row[2]
                code = pole_name.split("(")[1].rstrip(")") if "(" in str(pole_name) else str(pole_name)[0]
                poles[code] = float(value)
            coefficients[coeff_map[cell]] = poles
        elif cell == "APPRAISAL BASELINES":
            data_start = i + 3
            control_poles: dict[str, float] = {}
            certainty_poles: dict[str, float] = {}
            for offset in range(8):
                row = rows[data_start + offset]
                pole_name = row[1]
                code = pole_name.split("(")[1].rstrip(")") if "(" in str(pole_name) else str(pole_name)[0]
                control_poles[code] = float(row[2])
                certainty_poles[code] = float(row[3])
            coefficients["control"] = control_poles
            coefficients["certainty"] = certainty_poles
        i += 1

    return RawMBTIAxes(poles=coefficients)


def _parse_astrology(ws) -> RawAstrologyComponents:
    """Parse Sheet 2 (Astrology Components) into elements, modalities, tweaks."""
    rows = list(ws.iter_rows(values_only=True))
    elements: dict[str, dict[str, float]] = {}
    modalities: dict[str, dict[str, float]] = {}
    tweaks: dict[str, dict[str, float]] = {}

    section = None
    for row in rows:
        if row[0] == "ELEMENTS":
            section = "elements"
            continue
        elif row[0] == "MODALITIES":
            section = "modalities"
            continue
        elif row[0] and "SIGN-SPECIFIC TWEAKS" in str(row[0]):
            section = "tweaks"
            continue

        if row[0] is None:
            continue
        if row[0] in ("Element", "Modality", "Sign"):
            continue
        if all(v is None for v in row):
            continue
        if row[2] is None:
            continue

        name = row[0]
        values = {
            "assertiveness": float(row[2]),
            "susceptibility": float(row[3]),
            "rigidity": float(row[4]),
            "rumination": float(row[5]),
            "control": float(row[6]),
            "certainty": float(row[7]),
        }

        if section == "elements" and name in ("Fire", "Earth", "Air", "Water"):
            elements[name] = values
        elif section == "modalities" and name in ("Cardinal", "Fixed", "Mutable"):
            modalities[name] = values
        elif section == "tweaks":
            tweaks[name] = values

    return RawAstrologyComponents(elements=elements, modalities=modalities, tweaks=tweaks)


def _parse_mbti_matrix_components(ws) -> RawMBTIMatrixComponents:
    """Parse the MBTI Matrix Components sheet."""
    rows = list(ws.iter_rows(values_only=True))
    matrices: dict[str, list[list[float]]] = {}

    pole_names_to_codes = {
        "Extraversion (E)": "E", "Introversion (I)": "I",
        "Sensing (S)": "S", "Intuition (N)": "N",
        "Thinking (T)": "T", "Feeling (F)": "F",
        "Judging (J)": "J", "Perceiving (P)": "P",
    }

    i = 0
    while i < len(rows):
        cell = rows[i][0]
        if cell in pole_names_to_codes:
            code = pole_names_to_codes[cell]
            # Next row is column headers, then 5 rows of matrix data
            i += 2  # skip to first data row
            matrix = []
            for _ in range(5):
                row = rows[i]
                matrix.append([float(row[j]) for j in range(1, 6)])
                i += 1
            matrices[code] = matrix
            continue
        i += 1

    return RawMBTIMatrixComponents(matrices=matrices)


def _parse_mbti_baseline_components(ws) -> RawMBTIBaselineComponents:
    """Parse the MBTI Baseline Components sheet."""
    rows = list(ws.iter_rows(values_only=True))
    behavioral_baselines: dict[str, list[float]] = {}
    resting_moods: dict[str, list[float]] = {}

    section = None
    for row in rows:
        cell = row[0]
        if cell == "BEHAVIORAL BASELINES":
            section = "baselines"
            continue
        elif cell == "RESTING MOODS":
            section = "moods"
            continue
        elif cell == "Pole":
            continue  # skip header row

        if cell is None or cell in ("", " "):
            continue

        if section == "baselines" and cell in ("E", "I", "S", "N", "T", "F", "J", "P"):
            behavioral_baselines[cell] = [float(row[j]) for j in range(1, 6)]
        elif section == "moods" and cell in ("E", "I", "S", "N", "T", "F", "J", "P"):
            resting_moods[cell] = [float(row[j]) for j in range(1, 6)]

    return RawMBTIBaselineComponents(
        behavioral_baselines=behavioral_baselines,
        resting_moods=resting_moods,
    )


def _parse_astro_matrix_components(ws) -> RawAstroMatrixComponents:
    """Parse the Astro Matrix Components sheet."""
    rows = list(ws.iter_rows(values_only=True))

    all_elements = {"Fire", "Earth", "Air", "Water"}
    all_modalities = {"Cardinal", "Fixed", "Mutable"}
    all_signs = {
        "Aries", "Taurus", "Gemini", "Cancer", "Leo", "Virgo",
        "Libra", "Scorpio", "Sagittarius", "Capricorn", "Aquarius", "Pisces",
    }

    element_matrices: dict[str, list[list[float]]] = {}
    modality_matrices: dict[str, list[list[float]]] = {}
    tweak_matrices: dict[str, list[list[float]]] = {}

    section = None
    i = 0
    while i < len(rows):
        cell = rows[i][0]
        if cell == "ELEMENTS":
            section = "elements"
            i += 1
            continue
        elif cell == "MODALITIES":
            section = "modalities"
            i += 1
            continue
        elif cell == "SIGN-SPECIFIC TWEAKS":
            section = "tweaks"
            i += 1
            continue

        target_dict = None
        if section == "elements" and cell in all_elements:
            target_dict = element_matrices
        elif section == "modalities" and cell in all_modalities:
            target_dict = modality_matrices
        elif section == "tweaks" and cell in all_signs:
            target_dict = tweak_matrices

        if target_dict is not None:
            name = cell
            i += 2  # skip column headers, go to first data row
            matrix = []
            for _ in range(5):
                row = rows[i]
                matrix.append([float(row[j]) for j in range(1, 6)])
                i += 1
            target_dict[name] = matrix
            continue

        i += 1

    return RawAstroMatrixComponents(
        elements=element_matrices,
        modalities=modality_matrices,
        tweaks=tweak_matrices,
    )


def _parse_astro_baseline_components(ws) -> RawAstroBaselineComponents:
    """Parse the Astro Baseline Components sheet."""
    rows = list(ws.iter_rows(values_only=True))

    all_elements = {"Fire", "Earth", "Air", "Water"}
    all_modalities = {"Cardinal", "Fixed", "Mutable"}
    all_signs = {
        "Aries", "Taurus", "Gemini", "Cancer", "Leo", "Virgo",
        "Libra", "Scorpio", "Sagittarius", "Capricorn", "Aquarius", "Pisces",
    }

    element_baselines: dict[str, list[float]] = {}
    element_moods: dict[str, list[float]] = {}
    modality_baselines: dict[str, list[float]] = {}
    modality_moods: dict[str, list[float]] = {}
    tweak_baselines: dict[str, list[float]] = {}
    tweak_moods: dict[str, list[float]] = {}

    # Track which major section and subsection we're in
    major_section = None  # "BEHAVIORAL BASELINES" or "RESTING MOODS"
    sub_section = None  # "Elements", "Modalities", "Sign Tweaks"

    for row in rows:
        cell = row[0]

        if cell == "BEHAVIORAL BASELINES":
            major_section = "baselines"
            continue
        elif cell == "RESTING MOODS":
            major_section = "moods"
            continue
        elif cell == "Elements":
            sub_section = "elements"
            continue
        elif cell == "Modalities":
            sub_section = "modalities"
            continue
        elif cell == "Sign Tweaks":
            sub_section = "tweaks"
            continue

        if cell is None or cell == "Component":
            continue

        def _num(v):
            return 0.0 if v is None else float(v)

        values = [_num(row[j]) for j in range(1, 6)]

        if major_section == "baselines":
            if sub_section == "elements" and cell in all_elements:
                element_baselines[cell] = values
            elif sub_section == "modalities" and cell in all_modalities:
                modality_baselines[cell] = values
            elif sub_section == "tweaks" and cell in all_signs:
                tweak_baselines[cell] = values
        elif major_section == "moods":
            if sub_section == "elements" and cell in all_elements:
                element_moods[cell] = values
            elif sub_section == "modalities" and cell in all_modalities:
                modality_moods[cell] = values
            elif sub_section == "tweaks" and cell in all_signs:
                tweak_moods[cell] = values

    return RawAstroBaselineComponents(
        element_baselines=element_baselines,
        element_moods=element_moods,
        modality_baselines=modality_baselines,
        modality_moods=modality_moods,
        tweak_baselines=tweak_baselines,
        tweak_moods=tweak_moods,
    )


def _parse_reference_sheet(ws, key_column_name: str) -> dict[str, dict[str, float]]:
    """Parse a reference sheet (Sheet 3 or 4) into {name: {coeff: value}}."""
    rows = list(ws.iter_rows(values_only=True))
    result: dict[str, dict[str, float]] = {}

    header_row = None
    for i, row in enumerate(rows):
        if row[0] == key_column_name:
            header_row = i
            break

    if header_row is None:
        return result

    for row in rows[header_row + 1:]:
        name = row[0]
        if name is None:
            continue
        # Find the first numeric column
        data_start = next(
            (i for i, v in enumerate(row) if isinstance(v, (int, float))),
            None,
        )
        if data_start is None:
            continue
        result[name] = {
            "assertiveness": float(row[data_start]),
            "susceptibility": float(row[data_start + 1]),
            "rigidity": float(row[data_start + 2]),
            "rumination": float(row[data_start + 3]),
            "control": float(row[data_start + 4]),
            "certainty": float(row[data_start + 5]),
        }

    return result
